from openpyxl import load_workbook
from datetime import datetime
from tkinter import Tk, Text, TOP, BOTH, X, N, LEFT, BooleanVar 
from tkinter.ttk import Frame, Button, Style, Label, Entry


def write_into_workbook(values):
    loc = <user_file_location>

    wb = load_workbook(filename = loc)
    ws = wb['Weekly Sheet'] 

    now = datetime.now()
    msg = now.strftime('%m/%d/%Y')

    print(ws.max_row)
    col = 1
    row = ws.max_row + 1
  
    c = ws.cell(row, col)
    c.value =  msg
    d = ws.cell(row, col + 1) 
    d.value = values

    wb.save(loc)


class Example(Frame):

    def __init__(self):
        super().__init__()

        self.initUI()

    def centerWindow(self):

        w = 290
        h = 150

        sw = self.master.winfo_screenwidth()
        sh = self.master.winfo_screenheight()

        x = (sw - w)/2
        y = (sh - h)/2
        self.master.geometry('%dx%d+%d+%d' % (w, h, x, y))
A

    def initUI(self):
        self.master.title("Task Sheet")

def main():

    root = Tk()
    root.geometry("250x85+1250+650")
    app = Example()

    frame1 = Frame(root)
    frame1.pack(fill=BOTH)

    lbl1 = Label(frame1, text="Enter today's work")
    lbl1.pack(fill=X, padx=5, pady=5)

    frame2 = Frame(root)
    frame2.pack(fill=BOTH)

    entry1 = Entry(frame2)
    entry1.pack(fill=X, padx=5, expand=True)
    entry1.focus_set()

    def callback():
        print(entry1.get())
        print_values(entry1.get())
        write_into_workbook(entry1.get())

    frame3 = Frame(root)
    frame3.pack(fill=BOTH, expand=True)

    subbtn = Button(frame3, text="Done", command=callback)
    subbtn.place(x=4,y=5)

    root.mainloop()

if __name__ == '__main__':
    main()